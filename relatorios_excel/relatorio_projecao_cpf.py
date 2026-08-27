"""
relatorio_projecao_cpf.py
=========================
Gera relatório consolidado por distribuidora com:
  - Repique               : CPFs únicos ativos do arquivo repique_martina
  - Rodados c/ tit. ativa : CPFs únicos com CodigoRetorno 003 (excluindo CPFs do Repique)
  - Projeção              : estimativa de ATIVOs para o que ainda resta rodar
  - Total                 : soma das três classificações

Saída: relatorios_excel/relatorio_projecao_cpf.xlsx
"""

import re, json, os
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

ROOT   = Path(__file__).resolve().parent.parent
PASTA  = ROOT / 'dados' / 'filtrados_ativos'
OUTPUT = Path(__file__).parent / 'relatorio_projecao_cpf.xlsx'

DIST_ESTADO = {'PE': 'CELPE', 'BA': 'COELBA', 'RN': 'COSERN'}
DISTRIBUIDORAS = ['CELPE', 'COELBA', 'COSERN']

def norm_cpf(v):
    return re.sub(r'\D', '', str(v)).zfill(11)

def ler(p, sep=';', enc='utf-8'):
    for e in [enc, 'latin-1', 'utf-8']:
        try:
            df = pd.read_csv(p, dtype=str, sep=sep, encoding=e, on_bad_lines='skip')
            if len(df.columns) > 1:
                return df
        except Exception:
            pass
    return None

# ── 1. REPIQUE ────────────────────────────────────────────────────────────────
print('Carregando Repique...')
df_rep = ler(PASTA / 'repique_martina.csv', enc='latin-1')
df_rep['_cpf']  = df_rep['CPF/CNPJ'].apply(norm_cpf)
df_rep['_dist'] = df_rep['Estado'].map(DIST_ESTADO)
df_rep = df_rep[df_rep['_dist'].notna()]

# CPF único por distribuidora (um CPF pode ter 2 UCs em distribuidoras diferentes → conta nas 2)
repique_cpfs = set(df_rep['_cpf'])  # conjunto global para exclusão posterior

rep_counts = {}
for dist in DISTRIBUIDORAS:
    rep_counts[dist] = df_rep[df_rep['_dist'] == dist]['_cpf'].nunique()

print(f"  Repique: {sum(rep_counts.values()):,} registros | CPFs únicos por dist: {rep_counts}")

# ── 2. RODADOS COM TITULARIDADE ATIVA ─────────────────────────────────────────
print('Carregando Rodados com titularidade ativa...')

registros_ativos = []  # lista de (cpf_norm, dist)

# 2a. resultado_macro_local_completo.csv — filtrar CodigoRetorno 003
print('  resultado_macro_local_completo...')
df_res = pd.read_csv(ROOT / 'relatorios_excel' / 'resultado_macro_local_completo.csv', dtype=str, sep=';',
                     encoding='utf-8', on_bad_lines='skip')
for _, row in df_res.iterrows():
    try:
        j = json.loads(str(row['resposta']))
        if j.get('CodigoRetorno') == '003':
            emp = str(row.get('empresa', '')).strip().upper()
            if emp in DISTRIBUIDORAS:
                registros_ativos.append((norm_cpf(row['cpf']), emp))
    except Exception:
        pass
print(f'    {len(registros_ativos):,} ativos cod 003')

# 2b. titularidade_ligada_unificada_novo.csv — status ativo/ligada
print('  titularidade_ligada_unificada_novo...')
df_tit = ler(PASTA / 'titularidade_ligada_unificada_novo.csv')
df_tit = df_tit[df_tit['status'].str.lower().isin(['ativo', 'ligada'])]
for _, row in df_tit.iterrows():
    cpf = norm_cpf(row['cpf'])
    dist_raw = str(row.get('distribuidora', '')).strip().lower()
    for d in dist_raw.split('|'):
        d = d.strip().upper()
        if d in DISTRIBUIDORAS:
            registros_ativos.append((cpf, d))
print(f'    após titularidade: {len(registros_ativos):,}')

# 2c. NEO_BASE NOVA — já filtrado, DIST coluna
print('  NEO_BASE NOVA...')
df_neo = ler(PASTA / 'NEO_BASE NOVA_BA_PE_RN_02032026_REMOVIDODTS.csv')
for _, row in df_neo.iterrows():
    cpf = norm_cpf(row['cpf'])
    dist = str(row.get('DIST', '')).strip().upper()
    if dist in DISTRIBUIDORAS:
        registros_ativos.append((cpf, dist))
print(f'    após NEO_BASE: {len(registros_ativos):,}')

# 2d. pronto_para_importar — Estado → distribuidora
print('  pronto_para_importar...')
df_pront = ler(PASTA / 'pronto_para_importar_apenas_div_curvas_NEO_REMOVIDODTS.csv', enc='latin-1')
for _, row in df_pront.iterrows():
    cpf = norm_cpf(row['CPF/CNPJ'])
    dist = DIST_ESTADO.get(str(row.get('Estado', '')).strip().upper())
    if dist:
        registros_ativos.append((cpf, dist))
print(f'    após pronto_para_importar: {len(registros_ativos):,}')

# Deduplicar: por (cpf, dist), excluir CPFs presentes no repique
df_at = pd.DataFrame(registros_ativos, columns=['cpf', 'dist'])
df_at = df_at[~df_at['cpf'].isin(repique_cpfs)]
df_at = df_at.drop_duplicates(subset=['cpf', 'dist'])

rodados_counts = {}
for dist in DISTRIBUIDORAS:
    rodados_counts[dist] = int((df_at['dist'] == dist).sum())

print(f"  Rodados únicos (excl. repique): {sum(rodados_counts.values()):,} | por dist: {rodados_counts}")

# ── 3. PROJEÇÃO ───────────────────────────────────────────────────────────────
# Calcular taxa de aproveitamento POR ARQUIVO e aplicar sobre pendentes do mesmo arquivo
print('Calculando Projeção por arquivo...')

def norm_uc(v):
    return re.sub(r'\D', '', str(v)).lstrip('0') or '0'

DATA = ROOT / 'dados' / 'fornecedor2' / 'operacional'

ARQUIVOS = [
    ('06-04-2026', '06-04-2026/35K_20260402_CELP.csv',       'CELPE',   'cpf', 'uc'),
    ('06-04-2026', '06-04-2026/35K_20260402_COELBA.csv',     'COELBA',  'cpf', 'uc'),
    ('06-04-2026', '06-04-2026/35K_20260402_COSERN.csv',     'COSERN',  'cpf', 'uc'),
    ('06-04-2026', '06-04-2026/celpe_final_3103.xlsx',       'CELPE',   'cpf_consultado', 'uc'),
    ('13-04-2026', '13-04-2026/20260409_CELP_35K.csv',       'CELPE',   'cpf', 'uc'),
    ('13-04-2026', '13-04-2026/20260409_COELBA_35K.csv',     'COELBA',  'cpf', 'uc'),
    ('13-04-2026', '13-04-2026/20260409_COSERN_35K.csv',     'COSERN',  'cpf', 'uc'),
    ('16-04-2026', '16-04-2026/20260414_CELP_35K.csv',       'CELPE',   'cpf', 'uc'),
    ('16-04-2026', '16-04-2026/20260414_COELBA_35K.csv',     'COELBA',  'cpf', 'uc'),
    ('23-04-2026', '23-04-2026/20260422_CELP_35K.csv',       'CELPE',   'cpf', 'uc'),
    ('23-04-2026', '23-04-2026/coelba_15000.csv',            'COELBA',  'cpf', 'contract_account'),
    ('23-04-2026', '23-04-2026/cosern_15000.csv',            'COSERN',  'cpf', 'contract_account'),
    ('27-04-2026', '27-04-2026/celpe_15000_segundo_lote.csv',  'CELPE',  'cpf', 'contract_account'),
    ('27-04-2026', '27-04-2026/coelba_15000_segundo_lote.csv', 'COELBA', 'cpf', 'contract_account'),
    ('27-04-2026', '27-04-2026/cosern_15000_segundo_lote.csv', 'COSERN', 'cpf', 'contract_account'),
    ('04-05-2026', '04-05-2026/celpe_15000_terceiro_lote.csv',  'CELPE',  'cpf', 'contract_account'),
    ('04-05-2026', '04-05-2026/coelba_15000_terceiro_lote.csv', 'COELBA', 'cpf', 'contract_account'),
    ('04-05-2026', '04-05-2026/cosern_15000_terceiro_lote.csv', 'COSERN', 'cpf', 'contract_account'),
    ('04-05-2026', '04-05-2026/celpe_export_15000_20260502_153329.csv',  'CELPE',  'Cpf', 'Uc'),
    ('04-05-2026', '04-05-2026/coelba_export_15000_20260502_153437.csv', 'COELBA', 'Cpf', 'Uc'),
    ('04-05-2026', '04-05-2026/cosern_export_15000_20260502_153600.csv', 'COSERN', 'Cpf', 'Uc'),
    ('07-05-2026', '07-05-2026/CELPE_15000_20260507.xlsx',   'CELPE',   'documento', 'contrato'),
    ('07-05-2026', '07-05-2026/COELBA_15000_20260507.xlsx',  'COELBA',  'documento', 'contrato'),
    ('07-05-2026', '07-05-2026/COSERN_15000_20260507.xlsx',  'COSERN',  'documento', 'contrato'),
    ('11-05-2026', '11-05-2026/coelba_export_15000_20260509_194147.csv', 'COELBA', 'Cpf', 'Uc'),
    ('11-05-2026', '11-05-2026/cosern_export_15000_20260509_194217.csv', 'COSERN', 'Cpf', 'Uc'),
    ('11-05-2026', '11-05-2026/celpe_export_15000_20260509_194120.csv',  'CELPE',  'Cpf', 'Uc'),
    ('11-05-2026', '11-05-2026/coelba_export_15000_20260510_141945.csv', 'COELBA', 'Cpf', 'Uc'),
    ('11-05-2026', '11-05-2026/cosern_export_15000_20260510_142018.csv', 'COSERN', 'Cpf', 'Uc'),
    ('11-05-2026', '11-05-2026/celpe_export_15000_20260510_141914.csv',  'CELPE',  'Cpf', 'Uc'),
]

# Construir índice de resultados: (cpf_norm, uc_norm) -> CodigoRetorno
print('  Indexando resultados...')
df_full = pd.read_csv(ROOT / 'relatorios_excel' / 'resultado_macro_local_completo.csv',
                      dtype=str, sep=';', encoding='utf-8', on_bad_lines='skip')
resultado_idx = {}
for _, row in df_full.iterrows():
    try:
        j = json.loads(str(row['resposta']))
        cod = j.get('CodigoRetorno', '')
        k = (norm_cpf(row['cpf']), norm_uc(row['codigo cliente']))
        if k not in resultado_idx:
            resultado_idx[k] = cod
    except Exception:
        pass
print(f'  {len(resultado_idx):,} pares indexados')

# Por arquivo: calcular taxa e pendente
proj_por_dist = {d: 0.0 for d in DISTRIBUIDORAS}
vistos = set()

for data_lote, rel_path, dist, col_cpf, col_uc in ARQUIVOS:
    caminho = DATA / rel_path
    if not caminho.exists():
        continue
    try:
        if caminho.suffix == '.xlsx':
            df_src = pd.read_excel(caminho, dtype=str)
        else:
            with open(caminho, 'r', encoding='utf-8', errors='replace') as f:
                amostra = f.read(2048)
            sep = ';' if amostra.count(';') > amostra.count(',') else ','
            df_src = pd.read_csv(caminho, dtype=str, sep=sep, on_bad_lines='skip')
    except Exception:
        continue

    cols_lower = {c.lower(): c for c in df_src.columns}
    col_cpf_r = col_cpf if col_cpf in df_src.columns else cols_lower.get(col_cpf.lower())
    col_uc_r  = col_uc  if col_uc  in df_src.columns else cols_lower.get(col_uc.lower())
    if not col_cpf_r or not col_uc_r:
        continue

    df_src['_c'] = df_src[col_cpf_r].apply(norm_cpf)
    df_src['_u'] = df_src[col_uc_r].apply(norm_uc)
    df_src = df_src.dropna(subset=[col_cpf_r, col_uc_r])

    # inéditas (pares não vistos antes)
    ineditas = [(r['_c'], r['_u']) for _, r in df_src.iterrows() if (r['_c'], r['_u']) not in vistos]
    vistos |= set(zip(df_src['_c'], df_src['_u']))

    if not ineditas:
        continue

    ativos = sum(1 for k in ineditas if resultado_idx.get(k) == '003')
    inativos = sum(1 for k in ineditas if resultado_idx.get(k) in ('000','001','002','004','005','006'))
    pendente = len(ineditas) - ativos - inativos
    rodados = ativos + inativos
    taxa = ativos / rodados if rodados > 0 else 0
    proj_arq = pendente * taxa

    if proj_arq > 0 and dist in DISTRIBUIDORAS:
        proj_por_dist[dist] += proj_arq
        print(f'  {rel_path.split("/")[-1]:<50} dist={dist} taxa={taxa*100:.1f}% pend={pendente:,} proj={round(proj_arq):,}')

proj_counts = {d: round(v) for d, v in proj_por_dist.items()}
print(f'  Projeção final: {proj_counts}')

# ── 4. MONTAR TABELA FINAL ────────────────────────────────────────────────────
rows = []
for dist in DISTRIBUIDORAS:
    rows.append({'Distribuidora': dist, 'Classificação': 'Repique',                   'CPFs Únicos Ativos': rep_counts[dist]})
    rows.append({'Distribuidora': dist, 'Classificação': 'Rodados c/ Tit. Ativa',     'CPFs Únicos Ativos': rodados_counts[dist]})
    rows.append({'Distribuidora': dist, 'Classificação': 'Projeção (pendentes)',       'CPFs Únicos Ativos': proj_counts[dist]})
    total_dist = rep_counts[dist] + rodados_counts[dist] + proj_counts[dist]
    rows.append({'Distribuidora': dist, 'Classificação': 'TOTAL',                     'CPFs Únicos Ativos': total_dist})

# Total geral
for clf in ['Repique', 'Rodados c/ Tit. Ativa', 'Projeção (pendentes)', 'TOTAL']:
    v = sum(r['CPFs Únicos Ativos'] for r in rows if r['Classificação'] == clf)
    rows.append({'Distribuidora': 'TOTAL GERAL', 'Classificação': clf, 'CPFs Únicos Ativos': v})

df_out = pd.DataFrame(rows)

# ── 5. GERAR EXCEL ────────────────────────────────────────────────────────────
thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

CORES = {
    'Repique':               'FFF2CC',
    'Rodados c/ Tit. Ativa': 'E2EFDA',
    'Projeção (pendentes)':  'D9E1F2',
    'TOTAL':                 '2E75B6',
}

with pd.ExcelWriter(OUTPUT, engine='openpyxl') as writer:
    df_out.to_excel(writer, index=False, sheet_name='Projeção CPF')
    ws = writer.sheets['Projeção CPF']

    # Header
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[1].height = 28
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 22

    for row_idx, (_, row_data) in enumerate(df_out.iterrows(), start=2):
        clf   = row_data['Classificação']
        dist  = row_data['Distribuidora']
        is_total_row = clf == 'TOTAL' or dist == 'TOTAL GERAL'
        is_geral     = dist == 'TOTAL GERAL'

        if is_geral and clf == 'TOTAL':
            fill = PatternFill('solid', fgColor='1F4E79')
            font = Font(bold=True, color='FFFFFF', size=11)
        elif is_total_row:
            fill = PatternFill('solid', fgColor='2E75B6')
            font = Font(bold=True, color='FFFFFF', size=11)
        else:
            cor = CORES.get(clf, 'FFFFFF')
            fill = PatternFill('solid', fgColor=cor)
            font = Font(size=10)

        for cell in ws[row_idx]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

print(f'\nExcel salvo em: {OUTPUT}')
print('\nResumo:')
print(df_out.to_string(index=False))
