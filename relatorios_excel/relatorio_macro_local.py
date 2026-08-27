"""
relatorio_macro_local.py
========================
Gera Excel com resultados da macro local de validação de titularidade.

Lógica:
  - Para cada arquivo fonte (ordenado por data), calcula:
    * Total de pares CPF+UC no arquivo
    * Inéditas = pares que NÃO apareceram em nenhum arquivo de data anterior
    * Rodados  = inéditas que têm resultado em resultado_lote.csv
    * ATIVO / INATIVO / Pendente

Uso:
    python relatorios_excel/relatorio_macro_local.py

Saída: relatorios_excel/relatorio_macro_local.xlsx
"""

import sys, re, json
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / 'dados' / 'fornecedor2' / 'operacional'
RESULTADO = ROOT / 'relatorios_excel' / 'resultado_macro_local_completo.csv'
RESULTADO_01062026 = ROOT / 'macro_entrada_local' / 'dados' / 'resultado_consolidado_01062026.csv'

# Arquivos consolidados de rodagens anteriores (ordenados do mais antigo ao mais recente)
CONSOLIDADOS_EXTRAS = [
    ROOT / 'dados' / 'resultado_consolidado.csv',
    ROOT / 'macro_entrada_local' / 'dados' / 'resultado_lote_unificado.csv',
    ROOT / 'macro_entrada_local' / 'dados' / 'resultado_lote_merged_all.csv',
    ROOT / 'dados' / 'resultado_consolidado_maio2026.csv',
    ROOT / 'macro_entrada_local' / 'dados' / 'resultado_01062026.csv',
    ROOT / 'dados' / 'filtrados_ativos' / 'resultado_macro_local_completo.csv',
]

# ── filtro de mês no relatório ─────────────────────────────────────────────
# Exibe apenas datas >= A_PARTIR_DE (formato MM-YYYY).
# Os arquivos anteriores ainda são processados para o cálculo de "inéditas".
A_PARTIR_DE = (5, 2026)  # a partir de maio/2026

def exibir_data(data_lote):
    """Retorna True se a data_lote deve aparecer no relatório."""
    try:
        partes = data_lote.split('-')
        mes, ano = int(partes[1]), int(partes[2])
        return (ano, mes) >= (A_PARTIR_DE[1], A_PARTIR_DE[0])
    except Exception:
        return True

# ── helpers de normalização ──────────────────────────────────────────────────
def norm_cpf(v):
    return re.sub(r'\D', '', str(v)).zfill(11)

def norm_uc(v):
    return re.sub(r'\D', '', str(v)).lstrip('0') or '0'

def get_status(resp):
    try:
        cod = json.loads(str(resp)).get('CodigoRetorno', '')
        return 'ATIVO' if cod == '003' else 'INATIVO'
    except Exception:
        s = str(resp) if pd.notna(resp) else ''
        if 'TIMEOUT' in s: return 'TIMEOUT'
        if 'ERRO' in s or 'Error' in s: return 'ERRO'
        return s[:15] or '?'

# ── definição dos arquivos fonte ─────────────────────────────────────────────
# (data_lote, arquivo_relativo, empresa, col_cpf, col_uc)
# col_cpf / col_uc = None → detectar automaticamente
ARQUIVOS = [
    # 06-04-2026
    ('06-04-2026', '06-04-2026/35K_20260402_CELP.csv',       'celpe',   'cpf', 'uc'),
    ('06-04-2026', '06-04-2026/35K_20260402_COELBA.csv',     'coelba',  'cpf', 'uc'),
    ('06-04-2026', '06-04-2026/35K_20260402_COSERN.csv',     'cosern',  'cpf', 'uc'),
    ('06-04-2026', '06-04-2026/celpe_final_3103.xlsx',       'celpe',   'cpf_consultado', 'uc'),
    # 13-04-2026
    ('13-04-2026', '13-04-2026/20260409_CELP_35K.csv',       'celpe',   'cpf', 'uc'),
    ('13-04-2026', '13-04-2026/20260409_COELBA_35K.csv',     'coelba',  'cpf', 'uc'),
    ('13-04-2026', '13-04-2026/20260409_COSERN_35K.csv',     'cosern',  'cpf', 'uc'),
    # 16-04-2026
    ('16-04-2026', '16-04-2026/20260414_CELP_35K.csv',       'celpe',   'cpf', 'uc'),
    ('16-04-2026', '16-04-2026/20260414_COELBA_35K.csv',     'coelba',  'cpf', 'uc'),
    # 23-04-2026
    ('23-04-2026', '23-04-2026/20260422_CELP_35K.csv',       'celpe',   'cpf', 'uc'),
    ('23-04-2026', '23-04-2026/coelba_15000.csv',            'coelba',  'cpf', 'contract_account'),
    ('23-04-2026', '23-04-2026/cosern_15000.csv',            'cosern',  'cpf', 'contract_account'),
    # 27-04-2026
    ('27-04-2026', '27-04-2026/celpe_15000_segundo_lote.csv',  'celpe',  'cpf', 'contract_account'),
    ('27-04-2026', '27-04-2026/coelba_15000_segundo_lote.csv', 'coelba', 'cpf', 'contract_account'),
    ('27-04-2026', '27-04-2026/cosern_15000_segundo_lote.csv', 'cosern', 'cpf', 'contract_account'),
    # 04-05-2026 — terceiro lote
    ('04-05-2026', '04-05-2026/celpe_15000_terceiro_lote.csv',  'celpe',  'cpf', 'contract_account'),
    ('04-05-2026', '04-05-2026/coelba_15000_terceiro_lote.csv', 'coelba', 'cpf', 'contract_account'),
    ('04-05-2026', '04-05-2026/cosern_15000_terceiro_lote.csv', 'cosern', 'cpf', 'contract_account'),
    # 04-05-2026 — exports
    ('04-05-2026', '04-05-2026/celpe_export_15000_20260502_153329.csv',  'celpe',  'Cpf', 'Uc'),
    ('04-05-2026', '04-05-2026/coelba_export_15000_20260502_153437.csv', 'coelba', 'Cpf', 'Uc'),
    ('04-05-2026', '04-05-2026/cosern_export_15000_20260502_153600.csv', 'cosern', 'Cpf', 'Uc'),
    # 07-05-2026 — XLSX
    ('07-05-2026', '07-05-2026/CELPE_15000_20260507.xlsx',   'celpe',   'documento', 'contrato'),
    ('07-05-2026', '07-05-2026/COELBA_15000_20260507.xlsx',  'coelba',  'documento', 'contrato'),
    ('07-05-2026', '07-05-2026/COSERN_15000_20260507.xlsx',  'cosern',  'documento', 'contrato'),
    # 11-05-2026 — exports
    ('11-05-2026', '11-05-2026/coelba_export_15000_20260509_194147.csv', 'coelba', 'Cpf', 'Uc'),
    ('11-05-2026', '11-05-2026/cosern_export_15000_20260509_194217.csv', 'cosern', 'Cpf', 'Uc'),
    ('11-05-2026', '11-05-2026/celpe_export_15000_20260509_194120.csv',  'celpe',  'Cpf', 'Uc'),
    ('11-05-2026', '11-05-2026/coelba_export_15000_20260510_141945.csv', 'coelba', 'Cpf', 'Uc'),
    ('11-05-2026', '11-05-2026/cosern_export_15000_20260510_142018.csv', 'cosern', 'Cpf', 'Uc'),
    ('11-05-2026', '11-05-2026/celpe_export_15000_20260510_141914.csv',  'celpe',  'Cpf', 'Uc'),
    # 18-05-2026 — exports (arquivos físicos em 18-05-2026/)
    ('18-05-2026', '18-05-2026/celpe_export_30000_20260517_120942.csv',  'celpe',  'Cpf', 'Uc'),
    ('18-05-2026', '18-05-2026/coelba_export_30000_20260517_121051.csv', 'coelba', 'Cpf', 'Uc'),
    ('18-05-2026', '18-05-2026/cosern_export_30000_20260517_121127.csv', 'cosern', 'Cpf', 'Uc'),
    # 28-05-2026 — exports 30k (gerados em 24-05)
    ('28-05-2026', '28-05-2026/celpe_export_30000_20260524_205305.csv',  'celpe',  'Cpf', 'Uc'),
    ('28-05-2026', '28-05-2026/coelba_export_30000_20260524_205338.csv', 'coelba', 'Cpf', 'Uc'),
    ('28-05-2026', '28-05-2026/cosern_export_30000_20260524_205412.csv', 'cosern', 'Cpf', 'Uc'),
    # 29-05-2026 — exports 3.5k (gerados em 27-05)
    ('29-05-2026', '29-05-2026/celpe_export_3500_20260527_221126.csv',   'celpe',  'Cpf', 'Uc'),
    ('29-05-2026', '29-05-2026/coelba_export_3500_20260527_221211.csv',  'coelba', 'Cpf', 'Uc'),
    ('29-05-2026', '29-05-2026/cosern_export_3500_20260527_221255.csv',  'cosern', 'Cpf', 'Uc'),
    # 01-06-2026 — exports 30k (gerados em 31-05)
    ('01-06-2026', '01-06-2026/celpe_export_30000_20260531_203929.csv',  'celpe',  'Cpf', 'Uc'),
    ('01-06-2026', '01-06-2026/coelba_export_30000_20260531_204015.csv', 'coelba', 'Cpf', 'Uc'),
    ('01-06-2026', '01-06-2026/cosern_export_30000_20260531_204112.csv', 'cosern', 'Cpf', 'Uc'),
]

# ── carregar resultado — consolida todos os archives dos workers ──────────────
WORKERS_DIR = ROOT / 'workers'
worker_dirs = sorted(WORKERS_DIR.glob('worker_*'))

print(f'Consolidando resultado de {len(worker_dirs)} workers...')
resultado_idx = {}  # (cpf_norm, uc_norm) → status  (último resultado vence)

# 0) consolidados extras de rodagens anteriores (base histórica ampla)
for cons_path in CONSOLIDADOS_EXTRAS:
    if not cons_path.exists():
        continue
    try:
        with open(cons_path, 'r', encoding='utf-8', errors='replace') as fh:
            sample = fh.read(512)
        sep_c = ';' if sample.count(';') > sample.count(',') else ','
        df_c = pd.read_csv(cons_path, dtype=str, encoding='utf-8',
                           on_bad_lines='skip', sep=sep_c)
        if not {'cpf', 'codigo cliente', 'resposta'}.issubset(df_c.columns):
            continue
        df_c = df_c[df_c['resposta'].notna() & (df_c['resposta'].str.strip() != '')].copy()
        df_c['_c'] = df_c['cpf'].apply(norm_cpf)
        df_c['_u'] = df_c['codigo cliente'].apply(norm_uc)
        df_c['_s'] = df_c['resposta'].apply(get_status)
        antes = len(resultado_idx)
        for c, u, s in zip(df_c['_c'], df_c['_u'], df_c['_s']):
            resultado_idx[(c, u)] = s
        print(f'  {cons_path.name:<50} +{len(resultado_idx)-antes:>8,} novos | total: {len(resultado_idx):,}')
    except Exception as e:
        print(f'  [AVISO] {cons_path.name}: {e}')

# 1) resultado consolidado 01-06-2026 (workers recentes)
if RESULTADO_01062026.exists():
    try:
        df_jun = pd.read_csv(RESULTADO_01062026, dtype=str, encoding='utf-8', on_bad_lines='skip')
        if {'cpf', 'codigo cliente', 'resposta'}.issubset(df_jun.columns):
            df_jun = df_jun[df_jun['resposta'].notna() & (df_jun['resposta'].str.strip() != '')].copy()
            df_jun['_c'] = df_jun['cpf'].apply(norm_cpf)
            df_jun['_u'] = df_jun['codigo cliente'].apply(norm_uc)
            df_jun['_s'] = df_jun['resposta'].apply(get_status)
            antes = len(resultado_idx)
            for c, u, s in zip(df_jun['_c'], df_jun['_u'], df_jun['_s']):
                resultado_idx[(c, u)] = s
            print(f'  resultado_consolidado_01062026{" "*21} +{len(resultado_idx)-antes:>8,} novos | total: {len(resultado_idx):,}')
    except Exception as e:
        print(f'  [AVISO] Erro ao ler resultado consolidado 01-06: {e}')

# 2) base histórica (resultado_macro_local_completo.csv, se existir)
if RESULTADO.exists():
    try:
        df_base = pd.read_csv(RESULTADO, dtype=str, encoding='utf-8',
                              on_bad_lines='skip', sep=';')
        if {'cpf', 'codigo cliente', 'resposta'}.issubset(df_base.columns):
            df_base = df_base[df_base['resposta'].notna() & (df_base['resposta'].str.strip() != '')].copy()
            df_base['_c'] = df_base['cpf'].apply(norm_cpf)
            df_base['_u'] = df_base['codigo cliente'].apply(norm_uc)
            df_base['_s'] = df_base['resposta'].apply(get_status)
            for c, u, s in zip(df_base['_c'], df_base['_u'], df_base['_s']):
                resultado_idx[(c, u)] = s
            print(f'  Base histórica: {len(resultado_idx):,} pares carregados')
    except Exception as e:
        print(f'  [AVISO] Erro ao ler base histórica: {e}')

# 2) archives dos workers (sobrescrevem base histórica com dados mais recentes)
total_arquivos = 0
for wdir in worker_dirs:
    arq_dir  = wdir / 'dados' / 'arquivo'
    lote_csv = wdir / 'dados' / 'resultado_lote.csv'
    fontes   = []
    if arq_dir.exists():
        fontes += sorted(arq_dir.glob('*.csv'))
    if lote_csv.exists():
        fontes.append(lote_csv)

    for csv_path in fontes:
        try:
            # detectar separador (archives usam ',', resultado_lote usa ';')
            with open(csv_path, 'r', encoding='utf-8', errors='replace') as fh:
                sample = fh.read(1024)
            sep_arq = ';' if sample.count(';') > sample.count(',') else ','
            df_tmp = pd.read_csv(csv_path, dtype=str, encoding='utf-8',
                                 on_bad_lines='skip', sep=sep_arq)
            if not {'cpf', 'codigo cliente', 'resposta'}.issubset(df_tmp.columns):
                continue
            df_tmp = df_tmp[df_tmp['resposta'].notna() & (df_tmp['resposta'].str.strip() != '')].copy()
            df_tmp['_c'] = df_tmp['cpf'].apply(norm_cpf)
            df_tmp['_u'] = df_tmp['codigo cliente'].apply(norm_uc)
            df_tmp['_s'] = df_tmp['resposta'].apply(get_status)
            for c, u, s in zip(df_tmp['_c'], df_tmp['_u'], df_tmp['_s']):
                resultado_idx[(c, u)] = s
            total_arquivos += 1
        except Exception:
            pass

print(f'  {total_arquivos:,} arquivos lidos | {len(resultado_idx):,} pares únicos no resultado')

# ── processar cada arquivo ────────────────────────────────────────────────────
vistos_antes = set()   # pares CPF+UC já vistos (acumulativo por data)
resultados = []

data_anterior = None

for data_lote, rel_path, empresa, col_cpf, col_uc in ARQUIVOS:
    caminho = DATA / rel_path
    nome_arquivo = Path(rel_path).name
    obs = ''

    if not caminho.exists():
        print(f'  [AVISO] Não encontrado: {rel_path}')
        if exibir_data(data_lote):
            resultados.append({
                'Data Lote': data_lote, 'Arquivo': nome_arquivo, 'Distribuidora': empresa.upper(),
                'Total Arquivo': 0, 'CPF+UC Inéditas': 0, 'Rodados': 0, '% Rodado': 0.0,
                'ATIVO': 0, '% ATIVO': 0.0, 'INATIVO': 0, '% INATIVO': 0.0,
                'Erro/Timeout': 0, 'Pendente': 0, 'Observação': 'Arquivo não encontrado',
            })
        continue

    # Se mudou de data, congela os "vistos_antes" (não — acumula contínuo)
    # Inéditas = pares desse arquivo que não existem em nenhum arquivo anterior

    # Carregar arquivo
    try:
        if caminho.suffix == '.xlsx':
            df_src = pd.read_excel(caminho, dtype=str)
        else:
            # tentar detectar separador
            with open(caminho, 'r', encoding='utf-8', errors='replace') as f:
                amostra = f.read(2048)
            sep = ';' if amostra.count(';') > amostra.count(',') else ','
            df_src = pd.read_csv(caminho, dtype=str, sep=sep, on_bad_lines='skip')
    except Exception as e:
        obs = f'Erro ao ler: {e}'
        print(f'  [ERRO] {rel_path}: {e}')
        if exibir_data(data_lote):
            resultados.append({
                'Data Lote': data_lote, 'Arquivo': nome_arquivo, 'Distribuidora': empresa.upper(),
                'Total Arquivo': 0, 'CPF+UC Inéditas': 0, 'Rodados': 0, '% Rodado': 0.0,
                'ATIVO': 0, '% ATIVO': 0.0, 'INATIVO': 0, '% INATIVO': 0.0,
                'Pendente': 0, 'Observação': obs,
            })
        continue

    # normalizar nomes de coluna (case-insensitive fallback)
    cols_lower = {c.lower(): c for c in df_src.columns}
    col_cpf_real = col_cpf if col_cpf in df_src.columns else cols_lower.get(col_cpf.lower())
    col_uc_real  = col_uc  if col_uc  in df_src.columns else cols_lower.get(col_uc.lower())

    if not col_cpf_real or not col_uc_real:
        obs = f'Colunas não encontradas ({col_cpf}, {col_uc}) | disponíveis: {list(df_src.columns)[:6]}'
        print(f'  [AVISO] {rel_path}: {obs}')
        if not FILTRO_MES or FILTRO_MES in data_lote:
            resultados.append({
                'Data Lote': data_lote, 'Arquivo': nome_arquivo, 'Distribuidora': empresa.upper(),
                'Total Arquivo': len(df_src), 'CPF+UC Inéditas': 0, 'Rodados': 0, '% Rodado': 0.0,
                'ATIVO': 0, '% ATIVO': 0.0, 'INATIVO': 0, '% INATIVO': 0.0,
                'Pendente': 0, 'Observação': obs,
            })
        continue

    df_src['_c'] = df_src[col_cpf_real].apply(norm_cpf)
    df_src['_u'] = df_src[col_uc_real].apply(norm_uc)
    df_src = df_src.dropna(subset=[col_cpf_real, col_uc_real])

    total_arquivo = len(df_src)
    pares_arquivo = set(zip(df_src['_c'], df_src['_u']))

    # inéditas = pares deste arquivo que não foram vistos em arquivos anteriores
    ineditas_set = pares_arquivo - vistos_antes
    ineditas = len(ineditas_set)

    # atualiza acumulado
    vistos_antes |= pares_arquivo

    # resultados para as inéditas
    ativos = inativos = pendente = 0
    for par in ineditas_set:
        st = resultado_idx.get(par)
        if st is None:
            pendente += 1
        elif st == 'ATIVO':
            ativos += 1
        elif st == 'INATIVO':
            inativos += 1
        else:
            pendente += 1  # resposta inválida/erro = não conta como rodado

    rodados = ativos + inativos
    pct_rod = round(rodados / ineditas * 100, 1) if ineditas else 0.0
    pct_ati = round(ativos  / rodados * 100, 1)  if rodados  else 0.0
    pct_ina = round(inativos / rodados * 100, 1) if rodados  else 0.0

    if exibir_data(data_lote):
        resultados.append({
            'Data Lote': data_lote,
            'Arquivo': nome_arquivo,
            'Distribuidora': empresa.upper(),
            'Total Arquivo': total_arquivo,
            'CPF+UC Inéditas': ineditas,
            'Rodados': rodados,
            '% Rodado': pct_rod,
            'ATIVO': ativos,
            '% ATIVO': pct_ati,
            'INATIVO': inativos,
            '% INATIVO': pct_ina,
            'Pendente': pendente,
            'Observação': obs,
        })
        print(f'  [OK] {nome_arquivo:<52} total={total_arquivo:>6,}  inéditas={ineditas:>6,}  rodados={rodados:>6,}  ativo={ativos:>5,}  inativo={inativos:>6,}  pendente={pendente:>6,}')

# ── gerar Excel ───────────────────────────────────────────────────────────────
df = pd.DataFrame(resultados)
output = Path(__file__).parent / 'relatorio_macro_local.xlsx'

# Totais por data
totais = df.groupby('Data Lote')[['Total Arquivo','CPF+UC Inéditas','Rodados','ATIVO','INATIVO','Pendente']].sum().reset_index()
totais['Arquivo'] = '** TOTAL **'
totais['Distribuidora'] = ''
totais['% Rodado'] = (totais['Rodados'] / totais['CPF+UC Inéditas'] * 100).round(1)
totais['% ATIVO']  = (totais['ATIVO']   / totais['Rodados'] * 100).round(1).fillna(0)
totais['% INATIVO']= (totais['INATIVO'] / totais['Rodados'] * 100).round(1).fillna(0)
totais['Observação'] = ''
for col in ['% Rodado','% ATIVO','% INATIVO']:
    totais[col] = totais[col].fillna(0)

# intercalar totais após cada grupo de data
rows_final = []
for data, grupo in df.groupby('Data Lote', sort=False):
    rows_final.append(grupo)
    rows_final.append(totais[totais['Data Lote'] == data])
df_final = pd.concat(rows_final, ignore_index=True)

COL_ORDER = ['Data Lote','Arquivo','Distribuidora','Total Arquivo',
             'CPF+UC Inéditas','Rodados','% Rodado',
             'ATIVO','% ATIVO','INATIVO','% INATIVO','Pendente','Observação']
df_final = df_final[COL_ORDER]

with pd.ExcelWriter(output, engine='openpyxl') as writer:
    df_final.to_excel(writer, index=False, sheet_name='Macro Local')
    ws = writer.sheets['Macro Local']

    # header
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 35

    widths = {'A': 14, 'B': 50, 'C': 13, 'D': 14, 'E': 14, 'F': 10, 'G': 10,
              'H': 10, 'I': 10, 'J': 10, 'K': 10, 'L': 13, 'M': 40}
    for col_ltr, w in widths.items():
        ws.column_dimensions[col_ltr].width = w

    # paletas por data
    date_palettes = {
        '06-04-2026': ('D6DCE4', 'BDC5CF'),
        '13-04-2026': ('FCE4D6', 'F4CCBA'),
        '16-04-2026': ('FFF2CC', 'FFE699'),
        '23-04-2026': ('E2EFDA', 'C6E0B4'),
        '27-04-2026': ('D9E1F2', 'B4C6E7'),
        '04-05-2026': ('EBF3E8', 'D4E8CE'),
        '07-05-2026': ('FDE9D9', 'FAD0B6'),
        '11-05-2026': ('E8D5F5', 'D4AEF0'),
        '18-05-2026': ('DDEBF7', 'BDD7EE'),
        '28-05-2026': ('E2EFDA', 'C6E0B4'),
        '29-05-2026': ('FFF2CC', 'FFE699'),
        '01-06-2026': ('FCE4D6', 'F4CCBA'),
    }
    total_fill   = PatternFill('solid', fgColor='2E75B6')
    fill_obs     = PatternFill('solid', fgColor='FFF2CC')
    thin         = Side(style='thin', color='AAAAAA')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_fill   = PatternFill('solid', fgColor='70AD47')
    orange_fill  = PatternFill('solid', fgColor='ED7D31')

    lote_alt = {}
    for row_idx, (_, row_data) in enumerate(df_final.iterrows(), start=2):
        data_lote = row_data['Data Lote']
        arquivo   = row_data['Arquivo']
        is_total  = arquivo == '** TOTAL **'

        if is_total:
            fill = total_fill
            for cell in ws[row_idx]:
                cell.fill = fill
                cell.font = Font(bold=True, color='FFFFFF', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
        else:
            lote_alt[data_lote] = lote_alt.get(data_lote, 0) + 1
            alt = lote_alt[data_lote] % 2 == 0
            c1, c2 = date_palettes.get(data_lote, ('DDEEFF', 'BBDDFF'))
            fill = PatternFill('solid', fgColor=c2 if alt else c1)
            for cell in ws[row_idx]:
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # observação
            if row_data['Observação']:
                ws.cell(row_idx, 13).fill = fill_obs
                ws.cell(row_idx, 13).font = Font(italic=True, size=9, color='7F6000')
                ws.cell(row_idx, 13).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # ATIVO verde
            ativo_cell = ws.cell(row_idx, 8)
            if ativo_cell.value and int(ativo_cell.value) > 0:
                ativo_cell.fill = green_fill
                ativo_cell.font = Font(bold=True, color='FFFFFF', size=10)

            # Pendente laranja
            pend_cell = ws.cell(row_idx, 12)
            if pend_cell.value and int(pend_cell.value) > 0:
                pend_cell.fill = orange_fill
                pend_cell.font = Font(bold=True, color='FFFFFF', size=10)

    # borda no header
    for cell in ws[1]:
        cell.border = border

print(f'\nExcel salvo em: {output}')
print(f'\nResumo geral:')
print(df.groupby('Data Lote')[['CPF+UC Inéditas','Rodados','ATIVO','INATIVO','Pendente']].sum().to_string())
