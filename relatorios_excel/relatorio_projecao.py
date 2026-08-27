"""
relatorio_projecao.py
=====================
Relatório de projeção de ativos:

  1. Pré-22/04  → todos os registros de dados/filtrados_ativos (todos confirmados)
  2. Pós-22/04  → lê do relatorio_macro_local mais recente, filtra Data Lote > 22/04/2026
                  Ativos confirmados (cód. 003) + projeção dos pendentes pelo % do arquivo

Saída: relatorio_projecao_<timestamp>.xlsx
"""

import re
from pathlib import Path
from datetime import datetime, date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

ROOT      = Path(__file__).resolve().parent.parent
FILTRADOS = ROOT / 'dados' / 'filtrados_ativos' / 'titularidade_ligada_unificada_novo.csv'
REL_DIR   = Path(__file__).parent

TS     = datetime.now().strftime('%Y%m%d_%H%M%S')
output = REL_DIR / f'relatorio_projecao_{TS}.xlsx'

# ── encontrar relatorio_macro_local mais recente ──────────────────────────────
macro_files = sorted(REL_DIR.glob('relatorio_macro_local_*.xlsx'), reverse=True)
if not macro_files:
    # tenta o fixo
    macro_files = [REL_DIR / 'relatorio_macro_local.xlsx']
MACRO_XLSX = macro_files[0]
print(f'Usando: {MACRO_XLSX.name}')

# ── 1. Filtrados ativos pré-22/04 ────────────────────────────────────────────
print('Carregando filtrados_ativos...')
df_filt = pd.read_csv(FILTRADOS, dtype=str, on_bad_lines='skip', sep=None, engine='python')
ativos_pre        = len(df_filt)
ativos_pre_celpe  = int((df_filt['distribuidora'].str.lower() == 'celpe').sum())
ativos_pre_coelba = int((df_filt['distribuidora'].str.lower() == 'coelba').sum())
ativos_pre_cosern = int((df_filt['distribuidora'].str.lower() == 'cosern').sum())
print(f'  Ativos pré-22/04: {ativos_pre:,}  (celpe={ativos_pre_celpe:,}  coelba={ativos_pre_coelba:,}  cosern={ativos_pre_cosern:,})')

# ── 2. Ler relatorio_macro_local — apenas datas > 22/04/2026 ─────────────────
print(f'Carregando {MACRO_XLSX.name}...')
df_mac = pd.read_excel(MACRO_XLSX, dtype=str)

# remover linhas de total
df_mac = df_mac[df_mac['Arquivo'] != '** TOTAL **'].copy()

# converter Data Lote para date
def parse_data(v):
    try:
        return datetime.strptime(str(v).strip(), '%d-%m-%Y').date()
    except:
        return None

df_mac['_data'] = df_mac['Data Lote'].apply(parse_data)
corte = date(2026, 4, 22)
df_pos = df_mac[df_mac['_data'] > corte].copy()

# converter colunas numéricas
for col in ['CPF+UC Inéditas', 'Rodados', 'ATIVO', 'INATIVO', 'Pendente', '% ATIVO']:
    df_pos[col] = pd.to_numeric(df_pos[col], errors='coerce').fillna(0)

print(f'  Linhas pós-22/04: {len(df_pos)}  |  Lotes: {df_pos["Data Lote"].nunique()}')

# ── 3. Calcular projeção por arquivo ─────────────────────────────────────────
# % ATIVO já vem no Excel (ativos/rodados * 100).
# Para arquivos sem rodados, usa média do lote.
df_pos['_taxa'] = df_pos['% ATIVO'] / 100.0

# onde rodados == 0, preenche com média do lote
for dl in df_pos['Data Lote'].unique():
    mask_lote  = df_pos['Data Lote'] == dl
    mask_valid = mask_lote & (df_pos['Rodados'] > 0)
    if mask_valid.any():
        taxa_media = (df_pos.loc[mask_valid, 'ATIVO'].sum() /
                      df_pos.loc[mask_valid, 'Rodados'].sum())
    else:
        taxa_media = 0.0
    mask_sem = mask_lote & (df_pos['Rodados'] == 0)
    df_pos.loc[mask_sem, '_taxa'] = taxa_media

df_pos['Proj. Ativos']   = (df_pos['Pendente'] * df_pos['_taxa']).round().astype(int)
df_pos['Total Estimado'] = df_pos['ATIVO'].astype(int) + df_pos['Proj. Ativos']

total_conf_pos = int(df_pos['ATIVO'].sum())
total_proj_pos = int(df_pos['Proj. Ativos'].sum())

# ── 4. Gerar Excel ───────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = 'Projeção'

thin   = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
fill_hdr   = PatternFill('solid', fgColor='2E75B6')
fill_pre   = PatternFill('solid', fgColor='70AD47')
fill_tot   = PatternFill('solid', fgColor='1F4E79')
fill_proj  = PatternFill('solid', fgColor='FFE699')
fill_conf  = PatternFill('solid', fgColor='A9D18E')
font_hdr   = Font(bold=True, color='FFFFFF', size=11)
font_tot   = Font(bold=True, color='FFFFFF', size=12)
font_norm  = Font(size=10)
font_bold  = Font(bold=True, size=10)
align_c    = Alignment(horizontal='center', vertical='center')
align_l    = Alignment(horizontal='left',   vertical='center')

def wc(row, col, val='', fill=None, font=None, align=None, fmt=None):
    c = ws.cell(row, col, val)
    if fill:  c.fill   = fill
    if font:  c.font   = font
    if align: c.alignment = align
    c.border = border
    if fmt:   c.number_format = fmt
    return c

# Título
ws.merge_cells('A1:L1')
c = ws.cell(1, 1, '📊  PROJEÇÃO DE ATIVOS — VALIDAÇÃO DE TITULARIDADE')
c.fill = fill_tot; c.font = Font(bold=True, color='FFFFFF', size=14)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = border
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:L2')
c2 = ws.cell(2, 1, f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  Base: {MACRO_XLSX.name}')
c2.font = Font(italic=True, size=9, color='666666')
c2.alignment = align_l; c2.border = border

# ── Bloco pré-22/04 ──────────────────────────────────────────────────────────
row = 4
ws.merge_cells(f'A{row}:L{row}')
c = ws.cell(row, 1, '▶  ATIVOS PRÉ-22/04/2026  —  fonte: filtrados_ativos (todos confirmados)')
c.fill = fill_pre; c.font = font_hdr; c.alignment = align_l; c.border = border
row += 1

hdrs = ['Grupo', 'Distribuidora', 'Ativos Confirmados']
for i, h in enumerate(hdrs, 1):
    wc(row, i, h, fill=PatternFill('solid', fgColor='548235'), font=font_hdr, align=align_c)
row += 1

for dist, qtd in [('CELPE', ativos_pre_celpe), ('COELBA', ativos_pre_coelba),
                  ('COSERN', ativos_pre_cosern), ('TOTAL', ativos_pre)]:
    f = PatternFill('solid', fgColor='E2EFDA') if dist != 'TOTAL' else fill_pre
    fo = font_norm if dist != 'TOTAL' else font_hdr
    wc(row, 1, 'Pré-22/04', fill=f, font=fo, align=align_c)
    wc(row, 2, dist, fill=f, font=fo, align=align_c)
    wc(row, 3, qtd, fill=f, font=fo, align=align_c, fmt='#,##0')
    for col in range(4, 13): wc(row, col, '', fill=f)
    row += 1

# ── Bloco pós-22/04 ──────────────────────────────────────────────────────────
row += 1
ws.merge_cells(f'A{row}:L{row}')
c = ws.cell(row, 1, '▶  ATIVOS PÓS-22/04/2026  —  confirmados (cód. 003) + projeção pendentes')
c.fill = fill_hdr; c.font = font_hdr; c.alignment = align_l; c.border = border
row += 1

hdrs2 = ['Data Lote', 'Arquivo', 'Distribuidora', 'Inéditas', 'Rodados', '% Rodado',
         'Ativos Conf.', '% Ativo', 'Pendente', '% usado proj.', 'Proj. Ativos', 'Total Estimado']
for i, h in enumerate(hdrs2, 1):
    wc(row, i, h, fill=fill_hdr, font=font_hdr, align=align_c)
row += 1

date_palettes = {
    '23-04-2026': ('E2EFDA', 'C6E0B4'),
    '27-04-2026': ('D9E1F2', 'B4C6E7'),
    '04-05-2026': ('EBF3E8', 'D4E8CE'),
    '07-05-2026': ('FDE9D9', 'FAD0B6'),
    '11-05-2026': ('E8D5F5', 'D4AEF0'),
}
lote_alt = {}

for _, r in df_pos.iterrows():
    dl = r['Data Lote']
    lote_alt[dl] = lote_alt.get(dl, 0) + 1
    alt = lote_alt[dl] % 2 == 0
    c1, c2c = date_palettes.get(dl, ('DDEEFF', 'BBDDFF'))
    f = PatternFill('solid', fgColor=c2c if alt else c1)

    ativo_c  = int(r['ATIVO'])
    proj_v   = int(r['Proj. Ativos'])
    taxa     = float(r['_taxa'])
    vals = [dl, r['Arquivo'], r['Distribuidora'],
            int(r['CPF+UC Inéditas']), int(r['Rodados']),
            float(r['% Rodado']) / 100 if pd.notna(r['% Rodado']) else 0,
            ativo_c,
            float(r['% ATIVO']) / 100 if r['Rodados'] > 0 else '',
            int(r['Pendente']),
            taxa,
            proj_v,
            int(r['Total Estimado'])]
    fmts = [None, None, None, '#,##0', '#,##0', '0.0%',
            '#,##0', '0.0%', '#,##0', '0.0%', '#,##0', '#,##0']

    for col_i, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cc = wc(row, col_i, v, fill=f, font=font_norm, align=align_c, fmt=fmt)
        if col_i == 7 and ativo_c > 0:
            cc.fill = PatternFill('solid', fgColor='70AD47')
            cc.font = Font(bold=True, color='FFFFFF', size=10)
        if col_i == 11 and proj_v > 0:
            cc.fill = fill_proj
            cc.font = Font(bold=True, size=10)
    row += 1

# subtotal pós
f_sub = fill_hdr
wc(row, 1, 'SUBTOTAL PÓS-22/04', fill=f_sub, font=font_hdr, align=align_c)
for c in [2, 3, 5, 6, 8, 10]: wc(row, c, '', fill=f_sub)
wc(row, 4, int(df_pos['CPF+UC Inéditas'].sum()), fill=f_sub, font=font_hdr, align=align_c, fmt='#,##0')
wc(row, 5, int(df_pos['Rodados'].sum()),         fill=f_sub, font=font_hdr, align=align_c, fmt='#,##0')
wc(row, 7, total_conf_pos,  fill=f_sub, font=font_hdr, align=align_c, fmt='#,##0')
wc(row, 9, int(df_pos['Pendente'].sum()), fill=f_sub, font=font_hdr, align=align_c, fmt='#,##0')
wc(row, 11, total_proj_pos, fill=f_sub, font=font_hdr, align=align_c, fmt='#,##0')
wc(row, 12, total_conf_pos + total_proj_pos, fill=f_sub, font=font_hdr, align=align_c, fmt='#,##0')
row += 2

# ── Resumo final ─────────────────────────────────────────────────────────────
ws.merge_cells(f'A{row}:L{row}')
c = ws.cell(row, 1, '▶  RESUMO GERAL')
c.fill = fill_tot; c.font = Font(bold=True, color='FFFFFF', size=13)
c.alignment = align_l; c.border = border
ws.row_dimensions[row].height = 24
row += 1

total_conf  = ativos_pre + total_conf_pos
total_geral = total_conf + total_proj_pos

for lbl, val, ft in [
    ('Ativos pré-22/04 (confirmados — filtrados_ativos)',      ativos_pre,       fill_pre),
    ('Ativos pós-22/04 confirmados (cód. 003)',                total_conf_pos,   fill_conf),
    ('Subtotal confirmados',                                   total_conf,       PatternFill('solid', fgColor='548235')),
    ('Projeção dos pendentes',                                 total_proj_pos,   fill_proj),
    ('TOTAL GERAL ESTIMADO',                                   total_geral,      fill_tot),
]:
    ws.merge_cells(f'A{row}:K{row}')
    c = ws.cell(row, 1, lbl)
    fo = Font(bold=True, color='FFFFFF', size=11) if ft != fill_proj else Font(bold=True, size=11)
    c.fill = ft; c.font = fo; c.alignment = align_l; c.border = border
    vv = ws.cell(row, 12, val)
    vv.fill = ft; vv.font = fo; vv.alignment = align_c; vv.border = border
    vv.number_format = '#,##0'
    ws.row_dimensions[row].height = 22
    row += 1

# Larguras
for col, w in {'A': 14, 'B': 52, 'C': 14, 'D': 12, 'E': 12, 'F': 10,
               'G': 14, 'H': 10, 'I': 12, 'J': 13, 'K': 14, 'L': 16}.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A3'

wb.save(output)
print(f'\nExcel salvo em: {output.name}')
print(f'\n{"="*58}')
print(f'  Ativos pré-22/04 (confirmados):  {ativos_pre:>10,}')
print(f'  Ativos pós-22/04 conf. (003):    {total_conf_pos:>10,}')
print(f'  Subtotal confirmados:            {total_conf:>10,}')
print(f'  Projeção pendentes:              {total_proj_pos:>10,}')
print(f'  TOTAL GERAL ESTIMADO:            {total_geral:>10,}')
print(f'{"="*58}')

